import requests
import json
import datetime
import schedule
import time
from openpyxl import Workbook
from openpyxl import load_workbook

cep = '32677-530'
loadedWb = ''

def atualizarPlanilha():
    # tenta abrir a planilha
    try: 
        loadedWb = load_workbook('C:/Users/mathe/OneDrive/Área de Trabalho/Nova pasta/poc-vba-e-python-api-fetch/Controle.xlsx')
    # caso não exista, criar planilha nova
    except:  
        wb = Workbook()
        # pegar a planilha ativa
        ws = wb.active
        # buscar dados na api
        request = requests.get(f"https://cdn.apicep.com/file/apicep/{cep}.json")
        requestJSON = json.loads(request.content)
        #Cabeçalhos
        ws['A1'] = "Estado"
        ws['B1'] = "Cidade"
        ws['C1'] = "Bairro"
        ws['D1'] = "Endereço"
        ws['E1'] = "Data da atualização"

        row = (requestJSON["state"],requestJSON["city"],requestJSON["district"],requestJSON["address"], datetime.datetime.now())
        ws.append(row)
        # Salvar (Alterar caminho)
        wb.save("C:/Users/mathe/OneDrive/Área de Trabalho/Nova pasta/poc-vba-e-python-api-fetch/Controle.xlsx")
        # caso a planilha já exista, inserir nova linha
    else:
        # buscar dados na api
        request = requests.get(f"https://cdn.apicep.com/file/apicep/{cep}.json")
        requestJSON = json.loads(request.content)
        ws = loadedWb.active
        row = (requestJSON["state"],requestJSON["city"],requestJSON["district"],requestJSON["address"], datetime.datetime.now())
        ws.append(row)
        # Salvar (Alterar caminho)
        loadedWb.save("C:/Users/mathe/OneDrive/Área de Trabalho/Nova pasta/poc-vba-e-python-api-fetch/Controle.xlsx")

schedule.every(2).seconds.do(atualizarPlanilha)

atualizarPlanilha()

while True:
    schedule.run_pending()
    time.sleep(1)