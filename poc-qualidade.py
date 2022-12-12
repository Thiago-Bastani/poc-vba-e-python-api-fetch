import requests
import json
from openpyxl import Workbook
from openpyxl import load_workbook

cep = ''
loadedWb = ''
cep = input("Digite o CEP (xxxxx-xxx): ")
# tenta abrir a planilha
try: 
    loadedWb = load_workbook('test.xlsx')
    # caso não exista, criar planilha nova
except:  
    wb = Workbook()
    # grab the active worksheet
    ws = wb.active
    request = requests.get(f"https://cdn.apicep.com/file/apicep/{cep}.json")
    requestJSON = json.loads(request.content)
    #Cabeçalhos
    ws['A1'] = "Estado"
    ws['B1'] = "Cidade"
    ws['C1'] = "Bairro"
    ws['D1'] = "Endereço"

    row = (requestJSON["state"],requestJSON["city"],requestJSON["district"],requestJSON["address"])
    ws.append(row)
    # Save the file
    wb.save("test.xlsx")
    # caso a planilha já exista, inserir nova linha
else:
    request = requests.get(f"https://cdn.apicep.com/file/apicep/{cep}.json")
    requestJSON = json.loads(request.content)
    ws = loadedWb.active
    row = (requestJSON["state"],requestJSON["city"],requestJSON["district"],requestJSON["address"])
    ws.append(row)
    # Save the file
    loadedWb.save("test.xlsx")